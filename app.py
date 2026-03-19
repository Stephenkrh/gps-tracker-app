import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import time
import os
from math import radians, cos, sin, asin, sqrt
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime

st.set_page_config(page_title="GPS Tracker", layout="wide")
EXCEL_FILE = "GPS_Live_Data.xlsx"

# =========================
# SESSION STATE
# =========================
for key, val in {
    "tracking": False, "data": [], "kf_speed": 0.0,
    "tick": 0, "excel_ready": False, "start_time": None,
    "just_stopped": False,
    "prev_gps": None,       # ✅ track last processed GPS to avoid duplicates
}.items():
    if key not in st.session_state:
        st.session_state[key] = val

st.title("🚗 GPS Tracker")

# =========================
# SIDEBAR DEBUG
# =========================
st.sidebar.markdown("### 🔍 Debug Info")
st.sidebar.write(f"Rows in memory: **{len(st.session_state.data)}**")
st.sidebar.write(f"Tick: **{st.session_state.tick}**")
st.sidebar.write(f"Tracking: **{st.session_state.tracking}**")

# ✅ FIX 1: Read ALL query params — gps string split across params if too long
gps_lat     = st.query_params.get("lat",  None)
gps_lon     = st.query_params.get("lon",  None)
gps_acc     = st.query_params.get("acc",  None)
gps_spd     = st.query_params.get("spd",  None)
gps_hdg     = st.query_params.get("hdg",  None)
gps_alt     = st.query_params.get("alt",  None)
gps_ts      = st.query_params.get("ts",   None)
gps_err     = st.query_params.get("err",  None)

st.sidebar.write(f"Lat: `{gps_lat}` | Lon: `{gps_lon}`")
st.sidebar.write(f"Speed: `{gps_spd}` | Acc: `{gps_acc}`")
st.sidebar.write(f"Timestamp: `{gps_ts}`")
st.sidebar.write(f"Error: `{gps_err}`")

# =========================
# HELPERS
# =========================
def haversine(lat1, lon1, lat2, lon2):
    R = 6371
    dlat = radians(lat2 - lat1)
    dlon = radians(lon2 - lon1)
    a = (sin(dlat/2)**2 +
         cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon/2)**2)
    return 2 * R * asin(sqrt(a))

def kalman_filter(measured, prev_estimate, R=0.5):
    K = 1.0 / (1.0 + R)
    return prev_estimate + K * (measured - prev_estimate)

# =========================
# WRITE EXCEL ON STOP
# =========================
def write_excel_on_stop(data):
    if not data:
        return False
    df = pd.DataFrame(data)
    df["elapsed_sec"]       = (df["time"] - df["time"].iloc[0]).round(1)
    df["datetime"]          = (
        pd.to_datetime(df["time"], unit='s')
          .dt.tz_localize('UTC')
          .dt.tz_convert('Asia/Kolkata')
          .dt.strftime('%Y-%m-%d %H:%M:%S')
    )
    df["total_distance_km"] = df["distance_step"].cumsum().round(4)

    cols = ["elapsed_sec","datetime","lat","lon","accuracy_m",
            "speed","raw_speed","acc","heading","mode",
            "distance_step","total_distance_km"]
    df = df[[c for c in cols if c in df.columns]]

    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="GPS Trip Data")
        wb  = writer.book
        ws  = writer.sheets["GPS Trip Data"]
        hf  = PatternFill("solid", fgColor="1F4E79")
        hfn = Font(color="FFFFFF", bold=True, size=11)
        lf  = PatternFill("solid", fgColor="D6E4F0")

        for cell in ws[1]:
            cell.fill = hf
            cell.font = hfn
            cell.alignment = Alignment(horizontal="center")

        for i, row in enumerate(
                ws.iter_rows(min_row=2, max_row=ws.max_row), 1):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
                if i % 2 == 0:
                    cell.fill = lf

        for col in ws.columns:
            w = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = w + 4

        # Summary sheet
        ws2 = wb.create_sheet("Trip Summary")
        summary = [
            ("Trip Start",          df["datetime"].iloc[0]),
            ("Trip End",            df["datetime"].iloc[-1]),
            ("Total Duration (s)",  round(df["elapsed_sec"].iloc[-1], 1)),
            ("Total Distance (km)", round(df["total_distance_km"].iloc[-1], 3)),
            ("Avg Speed (km/h)",    round(df["speed"].mean(), 2)),
            ("Max Speed (km/h)",    round(df["speed"].max(), 2)),
            ("Max Accel (m/s²)",    round(df["acc"].max(), 4)),
            ("Total Rows",          len(df)),
        ]
        ws2.append(["Parameter", "Value"])
        for cell in ws2[1]:
            cell.fill = hf
            cell.font = Font(color="FFFFFF", bold=True)
        for row in summary:
            ws2.append(list(row))
        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 28
    return True

# =========================
# BUTTONS
# =========================
c1, c2, c3 = st.columns(3)

if c1.button("▶ Start Tracking"):
    st.session_state.data         = []
    st.session_state.kf_speed     = 0.0
    st.session_state.tick         = 0
    st.session_state.excel_ready  = False
    st.session_state.just_stopped = False
    st.session_state.prev_gps     = None
    st.session_state.start_time   = time.time()
    st.session_state.tracking     = True

if c2.button("⏹ Stop Tracking"):
    if st.session_state.tracking:
        st.session_state.tracking     = False
        st.session_state.just_stopped = True

if c3.button("🗑 Clear Data"):
    st.session_state.data         = []
    st.session_state.kf_speed     = 0.0
    st.session_state.tick         = 0
    st.session_state.excel_ready  = False
    st.session_state.just_stopped = False
    st.session_state.prev_gps     = None
    st.query_params.clear()
    if os.path.exists(EXCEL_FILE):
        os.remove(EXCEL_FILE)
    st.rerun()

# =========================
# WRITE EXCEL ON STOP
# =========================
if st.session_state.just_stopped:
    with st.spinner(f"💾 Saving {len(st.session_state.data)} rows..."):
        ok = write_excel_on_stop(st.session_state.data)
    st.session_state.excel_ready  = ok
    st.session_state.just_stopped = False
    if ok:
        st.success(
            f"✅ Saved {len(st.session_state.data)} rows to Excel"
        )
    else:
        st.warning("⚠️ No data to save — was tracking active?")

# =========================
# ✅ FIX 2: GPS COMPONENT — separate URL params per field
# Avoids truncation from long single param string
# =========================
components.html(
    """
    <!DOCTYPE html>
    <html>
    <body style="margin:0;background:#1e1e1e;padding:8px;
                 font-family:monospace;font-size:12px;color:#00ff88;">
    <div id="status">⏳ Requesting GPS...</div>

    <script>
    var lastTs = 0;

    function updateURLParams(params) {
        try {
            var url = new URL(window.parent.location.href);

            // ✅ FIX 3: Set each GPS field as a separate URL param
            // Avoids single long string truncation
            Object.keys(params).forEach(function(k) {
                url.searchParams.set(k, params[k]);
            });

            window.parent.history.replaceState(null, '', url.toString());
        } catch(e) {
            document.getElementById('status').innerText
                = '❌ URL update error: ' + e.message;
        }
    }

    function getGPS() {
        if (!navigator.geolocation) {
            document.getElementById('status').innerText
                = '❌ Geolocation not supported';
            return;
        }

        navigator.geolocation.getCurrentPosition(
            function(pos) {
                var ts = pos.timestamp;

                // ✅ FIX 4: Only update URL if timestamp changed
                // Prevents writing duplicate/stale GPS to URL
                if (ts === lastTs) {
                    document.getElementById('status').innerText
                        = '⚠️ Same GPS reading (no movement): ts=' + ts;
                    return;
                }
                lastTs = ts;

                // ✅ Full precision — separate params, no truncation
                var params = {
                    lat: pos.coords.latitude.toFixed(8),
                    lon: pos.coords.longitude.toFixed(8),
                    acc: (pos.coords.accuracy  || 0).toFixed(2),
                    spd: (pos.coords.speed     || 0).toFixed(4),
                    hdg: (pos.coords.heading   || 0).toFixed(2),
                    alt: (pos.coords.altitude  || 0).toFixed(2),
                    ts:  ts,
                    err: ''
                };

                updateURLParams(params);

                document.getElementById('status').innerText
                    = '✅ ' + params.lat + ', ' + params.lon
                    + ' | acc:' + params.acc + 'm'
                    + ' | spd:' + (params.spd*3.6).toFixed(1) + 'km/h'
                    + ' | ts:' + ts;
            },
            function(err) {
                var msgs = {
                    1: 'Permission denied',
                    2: 'Position unavailable',
                    3: 'Timeout — move outdoors'
                };
                updateURLParams({ err: err.code });
                document.getElementById('status').innerText
                    = '❌ Error ' + err.code
                    + ': ' + (msgs[err.code] || err.message);
            },
            {
                enableHighAccuracy: true,
                timeout:    15000,
                maximumAge: 2000
            }
        );
    }

    getGPS();
    setInterval(getGPS, 2000);
    </script>
    </body>
    </html>
    """,
    height=50,
)

# =========================
# PARSE GPS FROM SEPARATE URL PARAMS
# =========================
coords    = None
gps_error = None

if gps_err and gps_err != '':
    error_map = {
        "1": "Permission denied — allow location in browser",
        "2": "Position unavailable — check GPS signal",
        "3": "GPS timeout — move outdoors"
    }
    gps_error = error_map.get(str(gps_err), f"Error code {gps_err}")

elif gps_lat and gps_lon:
    try:
        coords = {
            "lat":       float(gps_lat),
            "lon":       float(gps_lon),
            "acc":       float(gps_acc  or 0),
            "gps_speed": float(gps_spd  or 0),
            "heading":   float(gps_hdg  or 0),
            "altitude":  float(gps_alt  or 0),
            "timestamp": float(gps_ts   or 0),
        }
    except Exception as e:
        gps_error = f"Parse error: {e}"

# =========================
# STATUS BAR
# =========================
status = st.empty()
if gps_error:
    status.error(f"❌ {gps_error}")
elif coords:
    elapsed = int(time.time() - st.session_state.start_time) \
              if st.session_state.start_time else 0
    status.success(
        f"✅ GPS Lock | "
        f"Lat: {coords['lat']:.8f} | "
        f"Lon: {coords['lon']:.8f} | "
        f"Acc: {coords['acc']:.1f}m | "
        f"Speed: {coords['gps_speed']*3.6:.1f} km/h | "
        f"Tick #{st.session_state.tick} | "
        f"Elapsed: {elapsed}s"
    )
else:
    status.warning("⏳ Waiting for GPS... Allow location in browser")

# =========================
# ✅ FIX 5: APPEND ROW — check timestamp to avoid duplicate rows
# =========================
if coords and st.session_state.tracking:

    curr_ts = coords["timestamp"]

    # ✅ Only append if GPS timestamp is new (not same reading twice)
    if curr_ts != st.session_state.prev_gps:
        st.session_state.prev_gps = curr_ts

        lat       = coords["lat"]
        lon       = coords["lon"]
        acc       = coords["acc"]
        gps_speed = coords["gps_speed"]
        heading   = coords["heading"]
        t         = time.time()

        if st.session_state.data:
            prev  = st.session_state.data[-1]
            dist  = haversine(prev["lat"], prev["lon"], lat, lon)
            dt    = max(t - prev["time"], 0.1)

            raw_speed    = (gps_speed * 3.6 if gps_speed > 0
                            else (dist / dt) * 3600)
            smooth_speed = kalman_filter(
                raw_speed, st.session_state.kf_speed
            )
            st.session_state.kf_speed = smooth_speed
            acc_val = ((smooth_speed - prev["speed"]) / 3.6) / dt
            mode    = ("Idle"    if smooth_speed < 2
                       else "Urban"   if smooth_speed < 40
                       else "Highway")
            new_row = {
                "time":          t,
                "lat":           round(lat, 8),   # ✅ full 8 decimal precision
                "lon":           round(lon, 8),
                "accuracy_m":    round(acc, 2),
                "speed":         round(smooth_speed, 2),
                "raw_speed":     round(raw_speed, 2),
                "acc":           round(acc_val, 4),
                "heading":       round(heading, 2),
                "mode":          mode,
                "distance_step": round(dist, 8),
            }
        else:
            new_row = {
                "time":          t,
                "lat":           round(lat, 8),
                "lon":           round(lon, 8),
                "accuracy_m":    round(acc, 2),
                "speed":         0.0,
                "raw_speed":     0.0,
                "acc":           0.0,
                "heading":       round(heading, 2),
                "mode":          "Idle",
                "distance_step": 0.0,
            }

        st.session_state.data.append(new_row)
        st.sidebar.success(
            f"✅ Row {len(st.session_state.data)} | "
            f"ts: {curr_ts}"
        )
    else:
        st.sidebar.warning(f"⚠️ Duplicate GPS ts={curr_ts} — skipped")

# =========================
# LIVE DASHBOARD
# =========================
if st.session_state.data:
    df     = pd.DataFrame(st.session_state.data)
    latest = df.iloc[-1]
    total  = df["distance_step"].sum()

    m1,m2,m3,m4,m5,m6 = st.columns(6)
    m1.metric("🚀 Speed",    f"{latest['speed']:.1f} km/h")
    m2.metric("⚡ Accel",    f"{latest['acc']:.3f} m/s²")
    m3.metric("🗺 Mode",      latest["mode"])
    m4.metric("📏 Distance", f"{total:.5f} km")
    m5.metric("🎯 Accuracy", f"{latest['accuracy_m']:.1f} m")
    m6.metric("📝 Rows",     len(df))

    t1, t2, t3 = st.tabs(["📈 Speed", "📉 Accel", "🗺 Map"])
    with t1: st.line_chart(df["speed"])
    with t2: st.line_chart(df["acc"])
    with t3: st.map(df[["lat","lon"]])

# =========================
# DOWNLOAD AFTER STOP
# =========================
st.divider()
if st.session_state.excel_ready and os.path.exists(EXCEL_FILE):
    with open(EXCEL_FILE, "rb") as f:
        st.download_button(
            "📥 Download Trip Excel", f.read(),
            f"GPS_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    st.caption(
        f"📊 {len(st.session_state.data)} rows — "
        f"1 per 2 seconds | 8 decimal GPS precision"
    )
else:
    st.info("▶ Start → drive → ⏹ Stop → download Excel")

# =========================
# AUTO REFRESH EVERY 2 SECONDS
# =========================
if st.session_state.tracking:
    st.session_state.tick += 1
    time.sleep(2)
    st.rerun()
